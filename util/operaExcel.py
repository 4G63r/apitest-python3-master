from openpyxl import load_workbook
import os

current_path = os.path.dirname(__file__)
case_path = '%s/cases.xlsx' % current_path.replace('util', 'testcase')


class OperaExcel:
    def __init__(self, file_name=None, sheet_name=None):
        if file_name and sheet_name:
            self.file_name = file_name
            self.sheet_name = sheet_name
        else:
            self.file_name = case_path
            self.sheet_name = '用例'
        self.ws = self.sheet

    @property
    def sheet(self):
        """获取worksheet对象"""
        self.wb = load_workbook(self.file_name)
        self.ws = self.wb[self.sheet_name]
        # sheets = wb.sheetnames  # 获取所有表格的名称
        return self.ws

    @property
    def sheet_rows(self):
        """获取有效行数"""
        return self.ws.max_row

    def get_cell(self, row, col):
        """获取单元格"""
        cell = self.ws.cell(row, col)
        return cell

    def get_cell_value(self, row, col):
        """获取单元格内容"""
        value = self.get_cell(row, col).value
        return value

    def write_data(self, row, col, data):
        """写入数据"""
        self.ws.cell(row, col).value = data

    def save_cell(self):
        """保存表格"""
        self.wb.save(self.file_name)

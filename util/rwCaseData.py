from util.operaExcel import OperaExcel
from util.operaJson import OperaJson
from openpyxl.styles import PatternFill


class GlobalVar:
    CASE_ID = 1
    CASE_NAME = 2
    IS_RUN = 3
    URL = 4
    METHOD = 5
    HEADERS = 6
    BODY_TYPE = 7
    BODY_CONTENT = 8
    DEPEND_CASE_ID = 9
    DEPEND_DATA = 10
    DEPEND_KEY = 11
    # EXPECT_STATUS_CODE = 12
    # EXPECT_RES_TIME = 13
    EXPECT_RES_VALUE = 12
    ACTUAL_RES = 13
    RES_TEXT = 14


class RWCaseData:
    def __init__(self):
        self.operaExcel = OperaExcel()

    @property
    def case_lines(self):
        """获取用例行数"""
        return self.operaExcel.sheet_rows

    def get_case_id(self, row):
        """获取用例id"""
        case_id = self.operaExcel.get_cell_value(row, col=GlobalVar.CASE_ID)
        if case_id:
            return case_id
        else:
            return None

    def get_case_name(self, row):
        """获取用例名称"""
        case_name = self.operaExcel.get_cell_value(row, col=GlobalVar.CASE_NAME)
        if case_name:
            return case_name
        else:
            return None

    def get_is_run(self, row):
        """获取是否执行"""
        is_run = self.operaExcel.get_cell_value(row, col=GlobalVar.IS_RUN)
        if is_run.upper() == 'YES':
            return True
        else:
            return False

    def get_url(self, row):
        """获取URL"""
        url = self.operaExcel.get_cell_value(row, col=GlobalVar.URL)
        if url:
            return url
        else:
            return None

    def get_request_method(self, row):
        """获取请求类型"""
        method = self.operaExcel.get_cell_value(row, col=GlobalVar.METHOD)
        if method.upper() == 'GET' or method.upper() == 'POST':
            return method
        else:
            return None

    def get_headers(self, row):
        """
        获取headers
        return字典
        """
        headers = self.operaExcel.get_cell_value(row, col=GlobalVar.HEADERS)
        if headers:
            header_dict = {}
            for i in headers.split('\n'):
                tmp = i.split(':')
                header_dict[tmp[0]] = tmp[1].strip()
        else:
            header_dict = {}
        return header_dict

    def get_body_type(self, row):
        """获取参数类型"""
        body_type = self.operaExcel.get_cell_value(row, col=GlobalVar.BODY_TYPE)
        if body_type:
            return body_type
        else:
            return None

    def get_body_content(self, row):
        """获取请求体内容"""
        body_content = self.operaExcel.get_cell_value(row, col=GlobalVar.BODY_CONTENT)
        if body_content == None:
            body_content = {}
        return body_content

    def get_depend_case_id(self, row):
        """获取依赖case id"""
        depend_case_id = self.operaExcel.get_cell_value(row, col=GlobalVar.DEPEND_CASE_ID)
        if depend_case_id:
            return depend_case_id
        else:
            return None

    def get_depend_data(self, row):
        """获取依赖数据"""
        depend_data = self.operaExcel.get_cell_value(row, col=GlobalVar.DEPEND_DATA)
        if depend_data:
            return depend_data
        else:
            return None

    def get_depend_key(self, row):
        """获取依赖的字段"""
        depend_key = self.operaExcel.get_cell_value(row, col=GlobalVar.DEPEND_KEY)
        if depend_key:
            return depend_key
        else:
            return None

    def get_expect_status_code(self, row):
        """预期响应状态码"""
        status_code = self.operaExcel.get_cell_value(row, col=GlobalVar.EXPECT_STATUS_CODE)
        if status_code:
            return status_code
        else:
            return None

    def get_expect_res_time(self, row):
        """预期响应时长(毫秒)"""
        res_time = self.operaExcel.get_cell_value(row, col=GlobalVar.EXPECT_RES_TIME)
        if res_time:
            return res_time
        else:
            return None

    def get_expect_res_value(self, row):
        """
        获取字段值
        return:响应是否包含字段值校验，非空字段值判断，字段长度校验，字段值校验
        """
        res_value = self.operaExcel.get_cell_value(row, col=GlobalVar.EXPECT_RES_VALUE)
        if res_value:
            check_values_contain = []
            check_null = []
            check_len = {}
            check_value_equal = {}
            if '\n' not in res_value:
                if ':' in res_value and '$' not in res_value:
                    check_values_contain.append(
                        '{}: {}'.format(res_value.split(':')[0], res_value.split(':')[1].strip()))
                elif ':' not in res_value:
                    check_null.append(res_value)
                elif res_value.count('$') == 1:
                    check_len[res_value.split(':')[0]] = int(res_value.split(':')[1].strip())
                elif res_value.count('$') == 2:
                    check_value_equal[res_value.split(':')[0]] = res_value.split(':')[1].strip()
                else:
                    raise Exception("字段值校验需要有':', 字段值非空校验只需要填写字段名, 序列长度校验需要一个'$', 字段和字段值验证需要两个'$'")
            else:
                for i in res_value.split('\n'):
                    if ':' in i and '$' not in i:
                        check_values_contain.append('{}: {}'.format(i.split(':')[0], i.split(':')[1].strip()))
                    elif ':' not in i:
                        check_null.append(i)
                    elif i.count('$') == 1:
                        check_len[i.split(':')[0]] = int(i.split(':')[1].strip())
                    elif i.count('$') == 2:
                        check_value_equal[i.split(':')[0]] = i.split(':')[1].strip()
                    else:
                        raise Exception("字段值校验需要有':', 字段值非空校验只需要填写字段名, 序列长度校验需要一个'$', 字段和字段值验证需要两个'$'")
            return check_values_contain, check_null, check_len, check_value_equal
        else:
            return None

    def get_actual_res(self, row):
        """获取实际结果"""
        res = self.operaExcel.get_cell_value(row, col=GlobalVar.ACTUAL_RES)
        return res

    def write_result(self, row, result, col=None):
        """写入测试结果并填充单元格"""
        if col == None:
            col = GlobalVar.ACTUAL_RES
        else:
            col = GlobalVar.RES_TEXT
        self.operaExcel.write_data(row, col, data=result)
        res = self.get_actual_res(row)
        self.set_cell_style(res, row)

    def get_body_by_json(self, row):
        """获取类型为字典格式的请求实体"""
        opera_json = OperaJson()
        res = self.get_body_content(row)
        data = opera_json.get_value(res)
        return data

    def color_fill(self, row, color):
        """单元格颜色填充"""
        cell = self.operaExcel.get_cell(row, col=GlobalVar.ACTUAL_RES)
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.fill = fill

    def set_cell_style(self, res, row):
        """
        设置单元格样式
        FF00FF00:绿色
        FFFF0000:红色
        """
        if res.upper() == 'PASS':
            self.color_fill(row, 'FF00FF00')
        elif res.upper() == 'FAIL':
            self.color_fill(row, 'FFFF0000')
        else:
            pass
        self.operaExcel.save_cell()

    def get_case_id_row(self, case_id):
        """获取case_id对应行号"""
        row = 0
        ws = self.operaExcel.ws
        for cell in ws['A']:
            if case_id == cell.value:
                row += 1
                break
            row += 1
        return row
